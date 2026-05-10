#!/usr/bin/env python3
"""
excel_to_data.py  —  Excel 여행플래너 → travelog/data.js 자동 변환
사용: python3 excel_to_data.py
"""
import openpyxl, json, re
from collections import OrderedDict
from pathlib import Path

SCRIPT_DIR  = Path(__file__).parent
EXCEL_PATH  = SCRIPT_DIR.parent / '2026_국토대장정_여행플래너.xlsx'
OUTPUT_PATH = SCRIPT_DIR / 'data.js'

DAY_REGION = {
    **{d: '수도권'    for d in range(1,  4)},
    **{d: '강원 영서' for d in range(4,  7)},
    **{d: '강원 영동' for d in range(7,  10)},
    **{d: '경북'      for d in range(10, 13)},
    **{d: '부산·경남' for d in range(13, 16)},
    **{d: '전남'      for d in range(16, 19)},
    **{d: '전북·충청' for d in range(19, 22)},
}
REGION_ORDER = ['수도권','강원 영서','강원 영동','경북','부산·경남','전남','전북·충청']

def norm_region(raw):
    r = str(raw).split('\n')[0].strip()
    for key, val in [('수도권','수도권'),('강원 영서','강원 영서'),
                     ('강원 영동','강원 영동'),('경북','경북'),
                     ('부산·경남','부산·경남'),('부산','부산·경남'),
                     ('통영','부산·경남'),('전남','전남'),('전북·충청','전북·충청')]:
        if r.startswith(key): return val
    return r

def get_category(meal, menu, name):
    m   = str(meal)
    txt = str(menu) + ' ' + str(name)
    if '디저트' in m or '간식' in m:
        return '디저트·간식'
    if any(k in txt for k in ['국밥','설렁탕','갈비탕','곰치국','미역국','해장국','순대국','물곰탕']):
        return '국밥·탕'
    if any(k in txt for k in ['전병','빈대떡','꼬막전','굴전','감자전','두부전','돼지고기전']):
        return '전·부침개'
    if any(k in txt for k in ['막국수','칼국수','수제비','옹심이','냉면','국수']):
        return '면·분식'
    if any(k in txt for k in ['갈비','삼겹','불고기','닭갈비','떡갈비','구이','간고등어']):
        return '고기·구이'
    return '백반·정식'

def map_query(name, addr):
    name_clean = re.sub(r'\s*[\(\(].*?[\)\)]','', name.split('\n')[0]).strip()
    parts = str(addr).split()
    city = parts[1].rstrip('시군구') if len(parts) >= 2 else ''
    return f"{name_clean} {city}".strip()

def s(v): return str(v).strip() if v else ''

def get_url(cell):
    """셀에서 URL 추출: 하이퍼링크 우선, 없으면 셀값이 URL인 경우"""
    if cell.hyperlink and cell.hyperlink.target:
        return cell.hyperlink.target.strip()
    val = str(cell.value or '').strip()
    if val.startswith('http://') or val.startswith('https://'):
        return val
    return None

def get_phone(cell):
    """셀에서 전화번호 추출"""
    val = str(cell.value or '').strip()
    m = re.search(r'0\d{1,2}-\d{3,4}-\d{4}', val)
    return m.group(0) if m else None

wb = openpyxl.load_workbook(EXCEL_PATH)

# ── SCHEDULE ──────────────────────────────────────────────────
ws = wb['일별일정표']
groups = OrderedDict((r, []) for r in REGION_ORDER)
for row in ws.iter_rows(min_row=3, values_only=True):
    d = row[0]
    if not isinstance(d, int) or not (1 <= d <= 21): continue
    groups[DAY_REGION[d]].append({
        'd': d, 'date': s(row[1]), 'dow': s(row[2]), 'loc': s(row[3]),
        'am': s(row[4]), 'pm': s(row[5]), 'eve': s(row[6]),
        'stay': s(row[7]), 'transport': s(row[8]),
    })
SCHEDULE = [{'region': r, 'days': days} for r, days in groups.items() if days]

# ── FOODS ──────────────────────────────────────────────────────
ws2 = wb['지역 찐맛집 추천']
FOODS = []
for row in ws2.iter_rows(min_row=4, values_only=True):
    if not row[0] or not row[2]: continue
    raw_r = str(row[0])
    if raw_r.startswith('==='): continue   # 구분선 스킵
    name  = row[2].split('\n')[0].strip() if row[2] and '\n' in str(row[2]) else s(row[2])
    if not name: continue
    meal  = s(row[1])
    menu  = s(row[3])
    addr  = s(row[5])
    FOODS.append({
        'region':   norm_region(raw_r),
        'meal':     meal,
        'category': get_category(meal, menu, name),
        'name':     name,
        'menu':     menu,
        'why':      s(row[4]),
        'addr':     addr,
        'map':      map_query(name, addr),
    })

# ── RESERVATIONS ──────────────────────────────────────────────
ws3 = wb['예약체크리스트']
RESERVATIONS = []
TYPE_MAP = {'숙박': '숙소', '교통': '교통', '체험': '체험'}
from datetime import date as _date
_START = _date(2026, 8, 3)

for row in ws3.iter_rows(min_row=3, values_only=False):
    구분 = s(row[0].value)
    if not 구분 or 구분 in ('구분', '여행 준비'): continue
    rtype = TYPE_MAP.get(구분)
    if not rtype: continue

    name     = s(row[1].value)
    date_str = s(row[2].value)
    plat     = s(row[3].value)
    note     = s(row[7].value)
    cell1, cell2 = row[8], row[9]

    # 하이퍼링크에서 URL 추출
    url1  = get_url(cell1)
    url2  = get_url(cell2)
    phone = get_phone(cell1) or get_phone(cell2)

    # 날짜 → Day 번호
    day_num = None
    m = re.search(r'(\d+)/(\d+)', date_str)
    if m:
        try:
            trip_date = _date(2026, int(m.group(1)), int(m.group(2)))
            day_num = (trip_date - _START).days + 1
        except Exception:
            pass

    rid = re.sub(r'[^a-z0-9]', '', name.lower())[:12] or f'r{len(RESERVATIONS)}'
    RESERVATIONS.append({
        'id':       rid,
        'type':     rtype,
        'name':     name,
        'day':      day_num,
        'date':     date_str,
        'desc':     plat + (' · ' + note if note else ''),
        'bookUrl':  url1,
        'bookUrl2': url2,
        'phone':    phone,
    })

# ── 출력 ──────────────────────────────────────────────────────
with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
    f.write('// AUTO-GENERATED — Excel에서 변환됨. 직접 수정 금지.\n\n')
    f.write(f'const SCHEDULE = {json.dumps(SCHEDULE, ensure_ascii=False, indent=2)};\n\n')
    f.write(f'const FOODS = {json.dumps(FOODS, ensure_ascii=False, indent=2)};\n\n')
    f.write(f'const RESERVATIONS = {json.dumps(RESERVATIONS, ensure_ascii=False, indent=2)};\n')

cats = {}
for f in FOODS:
    cats[f['category']] = cats.get(f['category'], 0) + 1
print(f'✅ data.js 생성: {len(SCHEDULE)}개 권역, {len(FOODS)}개 맛집, {len(RESERVATIONS)}개 예약')
print('카테고리별:', cats)
